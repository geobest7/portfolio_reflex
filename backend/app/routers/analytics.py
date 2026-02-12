from fastapi import APIRouter, Depends, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from datetime import datetime, timedelta
from io import BytesIO
from pydantic import BaseModel
from typing import Optional
from ..database import get_db
from ..models.visita import Visita
from ..utils.auth import get_current_admin_user

router = APIRouter()


# ==================== TRACKING (público, sin auth) ====================

class TrackingData(BaseModel):
    pagina: str
    referrer: str = ""
    user_agent: str = ""
    screen_width: Optional[int] = None
    screen_height: Optional[int] = None
    idioma: str = ""
    plataforma: str = ""


def detectar_dispositivo(user_agent: str, screen_width: Optional[int] = None) -> str:
    """Detectar dispositivo usando screen_width (más fiable) y user_agent como fallback"""
    if screen_width:
        if screen_width <= 768:
            return "movil"
        elif screen_width <= 1024:
            return "tablet"
        return "desktop"
    ua = user_agent.lower()
    if any(m in ua for m in ["mobile", "android", "iphone"]):
        return "movil"
    if "ipad" in ua or "tablet" in ua:
        return "tablet"
    return "desktop"


def detectar_navegador(user_agent: str) -> str:
    ua = user_agent.lower()
    if "edg" in ua:
        return "Edge"
    if "opr" in ua or "opera" in ua:
        return "Opera"
    if "chrome" in ua and "safari" in ua:
        return "Chrome"
    if "firefox" in ua:
        return "Firefox"
    if "safari" in ua:
        return "Safari"
    return "Otro"


def detectar_os(user_agent: str, plataforma: str = "") -> str:
    """Detectar sistema operativo"""
    p = plataforma.lower()
    if p:
        if "win" in p:
            return "Windows"
        if "mac" in p:
            return "macOS"
        if "linux" in p:
            return "Linux"
        if "iphone" in p or "ipad" in p:
            return "iOS"
        if "android" in p:
            return "Android"
    ua = user_agent.lower()
    if "iphone" in ua or "ipad" in ua:
        return "iOS"
    if "android" in ua:
        return "Android"
    if "windows" in ua:
        return "Windows"
    if "macintosh" in ua or "mac os" in ua:
        return "macOS"
    if "linux" in ua:
        return "Linux"
    return "Otro"


def anonimizar_ip(ip: str) -> str:
    if not ip:
        return ""
    partes = ip.split(".")
    if len(partes) == 4:
        partes[-1] = "0"
        return ".".join(partes)
    return ip


@router.delete("/limpiar-visitas")
def limpiar_visitas(db: Session = Depends(get_db), current_user=Depends(get_current_admin_user)):
    """Borrar todas las visitas (solo admin)"""
    count = db.query(Visita).count()
    db.query(Visita).delete()
    db.commit()
    return {"status": "ok", "borradas": count}


@router.post("/track")
async def registrar_visita(data: TrackingData, request: Request, db: Session = Depends(get_db)):
    """Endpoint público para registrar visitas desde el navegador del visitante"""
    ip_raw = request.headers.get("x-forwarded-for", "").split(",")[0].strip()
    if not ip_raw:
        ip_raw = request.client.host if request.client else ""
    
    visita = Visita(
        ip=anonimizar_ip(ip_raw),
        pagina=data.pagina or "/",
        metodo="GET",
        user_agent=(data.user_agent or "")[:500],
        referer=(data.referrer or "")[:500],
        pais="",
        ciudad="",
        dispositivo=detectar_dispositivo(data.user_agent, data.screen_width),
        navegador=detectar_navegador(data.user_agent),
        screen_width=data.screen_width,
        screen_height=data.screen_height,
        idioma=(data.idioma or "")[:10],
        plataforma=detectar_os(data.user_agent, data.plataforma),
    )
    db.add(visita)
    db.commit()
    
    return {"status": "ok"}


# ==================== CONSULTAS (requieren auth) ====================

@router.get("/resumen")
def resumen_analytics(
    dias: int = 30,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_admin_user)
):
    """Resumen general de analíticas"""
    fecha_desde = datetime.utcnow() - timedelta(days=dias)
    
    total = db.query(func.count(Visita.id)).filter(
        Visita.timestamp >= fecha_desde
    ).scalar()
    
    ips_unicas = db.query(func.count(func.distinct(Visita.ip))).filter(
        Visita.timestamp >= fecha_desde
    ).scalar()
    
    return {
        "total_visitas": total,
        "visitantes_unicos": ips_unicas,
        "periodo_dias": dias,
    }


@router.get("/paginas")
def paginas_mas_visitadas(
    dias: int = 30,
    limit: int = 10,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_admin_user)
):
    """Top páginas más visitadas"""
    fecha_desde = datetime.utcnow() - timedelta(days=dias)
    
    resultados = db.query(
        Visita.pagina,
        func.count(Visita.id).label("visitas")
    ).filter(
        Visita.timestamp >= fecha_desde
    ).group_by(Visita.pagina).order_by(desc("visitas")).limit(limit).all()
    
    return [{"pagina": r[0], "visitas": r[1]} for r in resultados]


@router.get("/dispositivos")
def distribucion_dispositivos(
    dias: int = 30,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_admin_user)
):
    """Distribución por tipo de dispositivo"""
    fecha_desde = datetime.utcnow() - timedelta(days=dias)
    
    resultados = db.query(
        Visita.dispositivo,
        func.count(Visita.id).label("total")
    ).filter(
        Visita.timestamp >= fecha_desde
    ).group_by(Visita.dispositivo).order_by(desc("total")).all()
    
    return [{"dispositivo": r[0], "total": r[1]} for r in resultados]


@router.get("/navegadores")
def distribucion_navegadores(
    dias: int = 30,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_admin_user)
):
    """Distribución por navegador"""
    fecha_desde = datetime.utcnow() - timedelta(days=dias)
    
    resultados = db.query(
        Visita.navegador,
        func.count(Visita.id).label("total")
    ).filter(
        Visita.timestamp >= fecha_desde
    ).group_by(Visita.navegador).order_by(desc("total")).all()
    
    return [{"navegador": r[0], "total": r[1]} for r in resultados]


@router.get("/plataformas")
def distribucion_plataformas(
    dias: int = 30,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_admin_user)
):
    """Distribución por sistema operativo"""
    fecha_desde = datetime.utcnow() - timedelta(days=dias)
    
    resultados = db.query(
        Visita.plataforma,
        func.count(Visita.id).label("total")
    ).filter(
        Visita.timestamp >= fecha_desde
    ).group_by(Visita.plataforma).order_by(desc("total")).all()
    
    return [{"plataforma": r[0] or "Desconocido", "total": r[1]} for r in resultados]


@router.get("/referrers")
def distribucion_referrers(
    dias: int = 30,
    limit: int = 10,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_admin_user)
):
    """De dónde vienen los visitantes"""
    fecha_desde = datetime.utcnow() - timedelta(days=dias)
    
    resultados = db.query(
        Visita.referer,
        func.count(Visita.id).label("total")
    ).filter(
        Visita.timestamp >= fecha_desde,
        Visita.referer != "",
        Visita.referer != None,
    ).group_by(Visita.referer).order_by(desc("total")).limit(limit).all()
    
    return [{"referrer": r[0], "total": r[1]} for r in resultados]


@router.get("/visitas-por-dia")
def visitas_por_dia(
    dias: int = 30,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_admin_user)
):
    """Visitas agrupadas por día"""
    fecha_desde = datetime.utcnow() - timedelta(days=dias)
    
    resultados = db.query(
        func.date(Visita.timestamp).label("fecha"),
        func.count(Visita.id).label("visitas")
    ).filter(
        Visita.timestamp >= fecha_desde
    ).group_by(func.date(Visita.timestamp)).order_by("fecha").all()
    
    return [{"fecha": str(r[0]), "visitas": r[1]} for r in resultados]


@router.get("/recientes")
def visitas_recientes(
    limit: int = 20,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_admin_user)
):
    """Últimas visitas registradas"""
    visitas = db.query(Visita).order_by(desc(Visita.timestamp)).limit(limit).all()
    
    return [
        {
            "id": v.id,
            "ip": v.ip,
            "pagina": v.pagina,
            "dispositivo": v.dispositivo,
            "navegador": v.navegador,
            "plataforma": v.plataforma or "",
            "referer": v.referer or "",
            "timestamp": v.timestamp.isoformat() if v.timestamp else "",
        }
        for v in visitas
    ]


@router.get("/export")
def exportar_excel(
    dias: int = 30,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_admin_user)
):
    """Exportar analíticas a Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    fecha_desde = datetime.utcnow() - timedelta(days=dias)
    
    wb = openpyxl.Workbook()
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="cccccc"),
        right=Side(style="thin", color="cccccc"),
        top=Side(style="thin", color="cccccc"),
        bottom=Side(style="thin", color="cccccc"),
    )
    
    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
    
    # === Hoja 1: Resumen ===
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    
    total = db.query(func.count(Visita.id)).filter(Visita.timestamp >= fecha_desde).scalar()
    unicos = db.query(func.count(func.distinct(Visita.ip))).filter(Visita.timestamp >= fecha_desde).scalar()
    
    ws_resumen.append(["Métrica", "Valor"])
    style_header(ws_resumen)
    ws_resumen.append(["Periodo", f"Últimos {dias} días"])
    ws_resumen.append(["Total Visitas", total])
    ws_resumen.append(["Visitantes Únicos", unicos])
    ws_resumen.column_dimensions["A"].width = 25
    ws_resumen.column_dimensions["B"].width = 20
    
    # === Hoja 2: Visitas por día ===
    ws_dias = wb.create_sheet("Visitas por Día")
    ws_dias.append(["Fecha", "Visitas"])
    style_header(ws_dias)
    
    por_dia = db.query(
        func.date(Visita.timestamp).label("fecha"),
        func.count(Visita.id).label("visitas")
    ).filter(Visita.timestamp >= fecha_desde).group_by(
        func.date(Visita.timestamp)
    ).order_by("fecha").all()
    
    for r in por_dia:
        ws_dias.append([str(r[0]), r[1]])
    ws_dias.column_dimensions["A"].width = 15
    ws_dias.column_dimensions["B"].width = 12
    
    # === Hoja 3: Todas las visitas ===
    ws_todas = wb.create_sheet("Todas las Visitas")
    ws_todas.append(["Fecha/Hora", "Página", "Dispositivo", "Navegador", "SO", "IP", "Referrer"])
    style_header(ws_todas)
    
    visitas = db.query(Visita).filter(
        Visita.timestamp >= fecha_desde
    ).order_by(desc(Visita.timestamp)).all()
    
    for v in visitas:
        ws_todas.append([
            v.timestamp.strftime("%Y-%m-%d %H:%M") if v.timestamp else "",
            v.pagina or "",
            v.dispositivo or "",
            v.navegador or "",
            v.plataforma or "",
            v.ip or "",
            v.referer or "",
        ])
    ws_todas.column_dimensions["A"].width = 18
    ws_todas.column_dimensions["B"].width = 30
    ws_todas.column_dimensions["C"].width = 12
    ws_todas.column_dimensions["D"].width = 12
    ws_todas.column_dimensions["E"].width = 12
    ws_todas.column_dimensions["F"].width = 16
    ws_todas.column_dimensions["G"].width = 30
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    fecha_str = datetime.utcnow().strftime("%Y%m%d")
    filename = f"analytics_{fecha_str}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
